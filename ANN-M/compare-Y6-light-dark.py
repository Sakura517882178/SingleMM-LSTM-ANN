"""
Y6 光照/暗态对照实验
训练同一 ANN 架构，分别使用 Y6-Light 和 Y6-Dark 的电流参数，
对比识别准确率，验证 Y6 光电导效应是器件工作的关键。

忆阻器权重直接使用 Y6 实测电流 (nA)，不再经过 log 域/电压转换。
"""
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader
import numpy as np
import os
import csv
import openpyxl

# ===========================
# Y6 参数加载：从 xlsx 提取电流 (nA)
# ===========================
def load_y6_params(xlsx_path, stable_after_ms=240.0):
    """从 Y6 I-V 数据提取电流范围 (nA)。
    Light: 50Hz 光脉冲调制，取 P5/P99.9 作为忆阻器工作范围。
    Dark:  无光调制，滤除仪器掉坑后取均值作为固定值 (无塑性)。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    time, current = [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            time.append(float(row[0]))
        if row[2] is not None:
            current.append(float(row[2]))
    t_arr = np.array(time)
    I_arr = np.array(current)
    I_stable = I_arr[t_arr > stable_after_ms]

    median_I = float(np.median(I_stable))
    # 暗态：大量仪器掉坑点 (I 接近 0)
    tail_frac = np.mean(I_stable < median_I * 0.1)
    if tail_frac > 0.05:
        I_clean = I_stable[I_stable > median_I * 0.1]
        I_val = float(np.mean(I_clean))
        return I_val, I_val  # 固定电流，无塑性
    else:
        return float(np.percentile(I_stable, 5)), float(np.percentile(I_stable, 99.9))


# ===========================
# Dataset
# ===========================
class SpeechFeatureDataset(torch.utils.data.Dataset):
    def __init__(self, npz_file):
        data = np.load(npz_file, allow_pickle=True)
        self.X = data["X"]
        self.y = data["y"]
        if self.y.min() < 0 or self.y.max() >= len(np.unique(self.y)):
            raise ValueError(f"标签范围不正确: min={self.y.min()}, max={self.y.max()}")
    def __len__(self): return len(self.X)
    def __getitem__(self, idx):
        return torch.tensor(self.X[idx], dtype=torch.float32), torch.tensor(self.y[idx], dtype=torch.long)

def collate_fn(batch):
    feats, labels = zip(*batch)
    return torch.stack(feats, dim=0), torch.stack(labels, dim=0)


# ===========================
# 忆阻器线性层（权重 = Y6 实测电流 nA）
# ===========================
class MemristorLinearI(nn.Module):
    def __init__(self, in_features, out_features, Imin, Imax, step=None, noise=0.0):
        super().__init__()
        self.Imin = Imin
        self.Imax = Imax
        self.step = step if step is not None else (Imax - Imin) / 100
        self.noise = noise
        self.I = nn.Parameter(torch.empty(out_features, in_features).uniform_(Imin, Imax))

    def forward(self, x):
        return F.linear(x, self.I)

    def update_memristor(self, lr=0.01):
        if self.I.grad is not None:
            delta = -lr * self.I.grad.data
            delta = torch.clamp(delta, -self.step, self.step)
            self.I.data += delta
            if self.noise > 0:
                self.I.data += self.noise * torch.randn_like(self.I)
            self.I.data = torch.clamp(self.I.data, self.Imin, self.Imax)
            self.I.grad.zero_()


# ===========================
# 忆阻器非线性层（替代 ReLU，权重 = Y6 实测电流 nA）
# ===========================
class MemristorNonlinear(nn.Module):
    def __init__(self, Imin, Imax, step=None, noise=0.0):
        super().__init__()
        self.Imin = Imin
        self.Imax = Imax
        self.step = step if step is not None else (Imax - Imin) / 100
        self.noise = noise
        self.I = nn.Parameter(torch.empty(1).uniform_(Imin, Imax))

    def forward(self, x):
        x = torch.clamp(x, min=0.0)
        return x * self.I

    def update_memristor(self, lr=0.001):
        if self.I.grad is not None:
            delta = -lr * self.I.grad.data
            delta = torch.clamp(delta, -self.step, self.step)
            self.I.data += delta
            if self.noise > 0:
                self.I.data += self.noise * torch.randn_like(self.I)
            self.I.data = torch.clamp(self.I.data, self.Imin, self.Imax)
            self.I.grad.zero_()


# ===========================
# 忆阻器版 ANN + 单向 LSTM + 可训练注意力
# ===========================
class MemristorANN_LSTM_Attn(nn.Module):
    def __init__(self, input_dim=64, hidden_dim=128, lstm_layers=1,
                 fc_hidden=128, num_classes=36, dropout=0.3,
                 Imin=0.4, Imax=32.0):
        super().__init__()
        self.lstm = nn.LSTM(input_dim, hidden_dim, lstm_layers,
                            batch_first=True, bidirectional=False)
        self.attn_weight = nn.Parameter(torch.Tensor(hidden_dim, 1))
        nn.init.xavier_uniform_(self.attn_weight)

        self.fc1 = MemristorLinearI(hidden_dim, fc_hidden, Imin, Imax)
        self.layernorm = nn.LayerNorm(fc_hidden)
        self.dropout = nn.Dropout(dropout)
        self.nonlinear = MemristorNonlinear(Imin, Imax)
        self.fc2 = MemristorLinearI(fc_hidden, num_classes, Imin, Imax)

    def forward(self, x):
        lstm_out, _ = self.lstm(x)
        attn_scores = torch.tanh(lstm_out @ self.attn_weight)
        attn_weights = torch.softmax(attn_scores, dim=1)
        pooled = torch.sum(lstm_out * attn_weights, dim=1)
        x = self.nonlinear(self.fc1(pooled))
        x = self.layernorm(x)
        x = self.dropout(x)
        logits = self.fc2(x)
        return logits

    def update_memristors(self, lr=0.01):
        for m in self.modules():
            if isinstance(m, (MemristorLinearI, MemristorNonlinear)):
                m.update_memristor(lr)


# ===========================
# 单次训练
# ===========================
def train_one(batch_size, num_epochs, patience, lr, memristor_lr, y6_file, label):
    device = "cuda" if torch.cuda.is_available() else "cpu"
    Imin, Imax = load_y6_params(y6_file)
    tag = "light" if "Dark" not in y6_file else "dark"

    train_dataset = SpeechFeatureDataset("train_features.npz")
    test_dataset = SpeechFeatureDataset("test_features.npz")
    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True, collate_fn=collate_fn)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)

    n_cls = len(np.unique(train_dataset.y))
    model = MemristorANN_LSTM_Attn(input_dim=64, hidden_dim=128, fc_hidden=128,
                                   num_classes=n_cls, Imin=Imin, Imax=Imax).to(device)

    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.AdamW(model.parameters(), lr=lr, weight_decay=1e-5)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.5, patience=5)

    best_acc = 0.0; epochs_no_improve = 0
    os.makedirs("checkpoints-AM-Uni", exist_ok=True)
    model_path = f"checkpoints-AM-Uni/best_memristor_lstm_attn_bs{batch_size}_{tag}.pth"
    csv_path = f"checkpoints-AM-Uni/log_memristor_lstm_attn_bs{batch_size}_{tag}.csv"

    with open(csv_path, mode='w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["epoch", "train_loss", "test_acc"])

    for epoch in range(num_epochs):
        model.train()
        running_loss = 0.0
        for X, y in train_loader:
            X, y = X.to(device), y.to(device)
            optimizer.zero_grad()
            out = model(X)
            loss = criterion(out, y)
            loss.backward()
            model.update_memristors(lr=memristor_lr)
            optimizer.step()
            running_loss += loss.item()

        avg_loss = running_loss / len(train_loader)

        model.eval()
        correct, total = 0, 0
        with torch.no_grad():
            for Xv, yv in test_loader:
                Xv, yv = Xv.to(device), yv.to(device)
                preds = model(Xv).argmax(dim=1)
                correct += (preds == yv).sum().item()
                total += yv.size(0)
        acc = correct / total
        scheduler.step(avg_loss)

        with open(csv_path, mode='a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([epoch+1, avg_loss, acc])

        if acc > best_acc:
            best_acc = acc
            epochs_no_improve = 0
            torch.save(model.state_dict(), model_path)
        else:
            epochs_no_improve += 1
        if epochs_no_improve >= patience:
            break

    print(f"  [{label}] bs={batch_size} best_acc={best_acc*100:.2f}%  "
          f"Imin={Imin:.3f} Imax={Imax:.3f} nA")
    return best_acc, Imin, Imax


# ===========================
# 主对比实验
# ===========================
def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--lr", type=float, default=0.003)
    parser.add_argument("--memristor_lr", type=float, default=0.001)
    parser.add_argument("--batch_sizes", type=int, nargs="+", default=[16, 32, 64])
    parser.add_argument("--epochs", type=int, default=200)
    parser.add_argument("--patience", type=int, default=30)
    args = parser.parse_args()

    light_file = "Y6-data.xlsx"
    dark_file = "Y6-Dark.xlsx"

    print("=" * 70)
    print("Y6 光照/暗态对照实验  (忆阻器权重 = Y6 实测电流 nA)")
    print("=" * 70)

    # 显示参数
    Il_min, Il_max = load_y6_params(light_file)
    Id_min, Id_max = load_y6_params(dark_file)
    print(f"\nLight: I=[{Il_min:.3f}, {Il_max:.3f}] nA")
    print(f"Dark:  I=[{Id_min:.4f}, {Id_max:.4f}] nA")
    print(f"光暗电流比: {Il_max / max(Id_max, 1e-9):.0f}x")
    print()

    results = {"Light": [], "Dark": []}

    for bs in args.batch_sizes:
        print(f"--- Batch size = {bs} ---")
        for y6_file, label in [(light_file, "Light"), (dark_file, "Dark")]:
            acc, imin, imax = train_one(bs, args.epochs, args.patience,
                                        args.lr, args.memristor_lr, y6_file, label)
            results[label].append(acc)

    # 输出汇总
    print("\n" + "=" * 70)
    print("对照实验结果汇总")
    print("=" * 70)
    print(f"{'Batch':>10s}  {'Light Acc':>12s}  {'Dark Acc':>12s}  {'Delta':>10s}")
    print("-" * 50)
    for i, bs in enumerate(args.batch_sizes):
        la = results["Light"][i] * 100
        da = results["Dark"][i] * 100
        print(f"{bs:>10d}  {la:>11.2f}%  {da:>11.2f}%  {la-da:>9.2f}%")
    print("-" * 50)
    light_best = max(max(results["Light"]), 0) * 100
    dark_best = max(max(results["Dark"]), 0) * 100
    print(f"{'Best':>10s}  {light_best:>11.2f}%  {dark_best:>11.2f}%  {light_best-dark_best:>9.2f}%")


if __name__ == "__main__":
    main()
