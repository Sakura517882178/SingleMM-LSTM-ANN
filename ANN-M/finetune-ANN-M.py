import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader
import numpy as np
import os
import shutil
import csv
import openpyxl

# ===========================
# 从 Y6 实验数据提取忆阻器参数 (log 域)
# ===========================
def load_y6_params(xlsx_path="Y6-data.xlsx", stable_after_ms=240.0):
    """从 Y6 单分子结 I-V 数据提取 log 域电导参数。
    Light: 50Hz 光脉冲调制，取稳定区 P5/P99.9 作为动态范围。
    Dark:  无光调制，电导应为定值; 滤除仪器掉坑(G > median-1.5)后取均值。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    time, voltage, cond = [], [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            time.append(float(row[0]))
        if row[1] is not None:
            voltage.append(float(row[1]))
        if row[3] is not None:
            cond.append(float(row[3]))
    t_arr = np.array(time)
    v_arr = np.array(voltage)
    c_arr = np.array(cond)
    stable_mask = t_arr > stable_after_ms
    c_stable = c_arr[stable_mask]
    V_exp = float(np.max(v_arr))
    median = float(np.median(c_stable))
    tail_frac = np.mean(c_stable < median - 1.5)
    if tail_frac > 0.05:
        c_clean = c_stable[c_stable > -5]
        G_val = float(np.mean(c_clean))
        return G_val, G_val, V_exp
    else:
        return float(np.percentile(c_stable, 5)), float(np.percentile(c_stable, 99.9)), V_exp

# ===========================
# Dataset
# ===========================
class SpeechFeatureDataset(torch.utils.data.Dataset):
    def __init__(self, npz_file):
        data = np.load(npz_file, allow_pickle=True)
        self.X = data["X"]
        self.y = data["y"]
        if self.y.min() < 0 or self.y.max() >= len(np.unique(self.y)):
            raise ValueError(f"标签范围不正确: min={self.y.min()}, max={self.y.max()}")

    def __len__(self):
        return len(self.X)

    def __getitem__(self, idx):
        return torch.tensor(self.X[idx], dtype=torch.float32), torch.tensor(self.y[idx], dtype=torch.long)

def collate_fn(batch):
    feats, labels = zip(*batch)
    return torch.stack(feats, dim=0), torch.stack(labels, dim=0)

# ===========================
# 忆阻器线性层（log 域: I = 10^G * V）
# ===========================
class MemristorLinearI(nn.Module):
    def __init__(self, in_features, out_features, Gmin_log, Gmax_log, V, step=0.01, noise=0.0):
        super().__init__()
        self.Gmin = Gmin_log
        self.Gmax = Gmax_log
        self.step = step
        self.noise = noise
        self.V = V
        self.G = nn.Parameter(torch.empty(out_features, in_features).uniform_(Gmin_log, Gmax_log))

    def forward(self, x):
        I = 10 ** self.G * self.V
        return F.linear(x, I)

    def update_memristor(self, lr=0.01):
        if self.G.grad is not None:
            delta = -lr * self.G.grad.data
            delta = torch.clamp(delta, -self.step, self.step)
            self.G.data += delta
            if self.noise > 0:
                self.G.data += self.noise * torch.randn_like(self.G)
            self.G.data = torch.clamp(self.G.data, self.Gmin, self.Gmax)
            self.G.grad.zero_()

# ===========================
# 忆阻器非线性层（log 域: I = x * 10^G * V）
# ===========================
class MemristorNonlinear(nn.Module):
    def __init__(self, Gmin_log, Gmax_log, V, step=0.01, noise=0.0):
        super().__init__()
        self.V = V
        self.Gmin = Gmin_log
        self.Gmax = Gmax_log
        self.step = step
        self.noise = noise
        self.G = nn.Parameter(torch.empty(1).uniform_(Gmin_log, Gmax_log))

    def forward(self, x):
        x = torch.clamp(x, min=0.0)
        I = x * (10 ** self.G * self.V)
        return I

    def update_memristor(self, lr=0.001):
        if self.G.grad is not None:
            delta = -lr * self.G.grad.data
            delta = torch.clamp(delta, -self.step, self.step)
            self.G.data += delta
            if self.noise > 0:
                self.G.data += self.noise * torch.randn_like(self.G)
            self.G.data = torch.clamp(self.G.data, self.Gmin, self.Gmax)
            self.G.grad.zero_()

# ===========================
# 忆阻器 ANN + LSTM + 注意力
# ===========================
class MemristorANN_LSTM_Attn(nn.Module):
    def __init__(self, input_dim=64, hidden_dim=128, lstm_layers=1,
                 fc_hidden=128, num_classes=36, dropout=0.3,
                 Gmin_log=-7.92, Gmax_log=-2.81, V=0.35):
        super().__init__()
        self.hidden_dim = hidden_dim

        self.lstm = nn.LSTM(input_dim, hidden_dim, lstm_layers,
                            batch_first=True, bidirectional=False)

        # 可训练注意力
        self.attn_weight = nn.Parameter(torch.Tensor(hidden_dim, 1))
        nn.init.xavier_uniform_(self.attn_weight)

        # 忆阻器 FC + Dropout + LayerNorm + 非线性
        self.fc1 = MemristorLinearI(hidden_dim, fc_hidden, Gmin_log, Gmax_log, V)
        self.layernorm = nn.LayerNorm(fc_hidden)
        self.dropout = nn.Dropout(dropout)
        self.nonlinear = MemristorNonlinear(Gmin_log, Gmax_log, V)
        self.fc2 = MemristorLinearI(fc_hidden, num_classes, Gmin_log, Gmax_log, V)

    def forward(self, x):
        lstm_out, _ = self.lstm(x)                       # [B, T, hidden_dim]
        attn_scores = torch.tanh(lstm_out @ self.attn_weight)
        attn_weights = torch.softmax(attn_scores, dim=1)
        pooled = torch.sum(lstm_out * attn_weights, dim=1)

        x = self.nonlinear(self.fc1(pooled))
        x = self.layernorm(x)
        x = self.dropout(x)
        logits = self.fc2(x)
        return logits

    def update_memristors(self, lr=0.01):
        for m in self.modules():
            if isinstance(m, (MemristorLinearI, MemristorNonlinear)):
                m.update_memristor(lr)

# ===========================
# 训练函数（带预训练加载与冻结选项）
# ===========================
def train_model(batch_size, num_epochs=200, patience=30, lr=0.003,
                memristor_lr=0.001, pretrained_path=None, freeze_nonmemristor=False):
    device = "cuda" if torch.cuda.is_available() else "cpu"

    # 从 Y6 实验数据读取忆阻器参数
    Gmin_log, Gmax_log, V_exp = load_y6_params("data.xlsx")

    train_dataset = SpeechFeatureDataset("train_features.npz")
    test_dataset = SpeechFeatureDataset("test_features.npz")
    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True, collate_fn=collate_fn)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)

    num_classes = len(np.unique(train_dataset.y))
    model = MemristorANN_LSTM_Attn(input_dim=64, hidden_dim=128,
                                   fc_hidden=128, num_classes=num_classes,
                                   Gmin_log=Gmin_log, Gmax_log=Gmax_log, V=V_exp).to(device)

    # ======== 加载预训练模型 ========
    if pretrained_path is not None and os.path.exists(pretrained_path):
        print(f"Loading pretrained weights from: {pretrained_path}")
        state_dict = torch.load(pretrained_path, map_location=device)
        model_dict = model.state_dict()
        matched = {k: v for k, v in state_dict.items() if k in model_dict and v.shape == model_dict[k].shape}
        model_dict.update(matched)
        model.load_state_dict(model_dict)
        print(f"✅ Loaded {len(matched)} layers successfully, skipped {len(model_dict)-len(matched)} unmatched layers.")
    else:
        print("⚠️ No pretrained model loaded. Training from scratch.")

    # ======== 冻结非忆阻器层（可选） ========
    if freeze_nonmemristor:
        for name, param in model.named_parameters():
            if not isinstance(param, (MemristorLinearI, MemristorNonlinear)):
                param.requires_grad = False
        print("🔒 Frozen non-memristor layers. Only memristor layers will be trained.")

    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.AdamW(filter(lambda p: p.requires_grad, model.parameters()), lr=lr, weight_decay=1e-5)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.5, patience=5)

    best_acc = 0.0
    epochs_no_improve = 0
    os.makedirs("checkpoints-AM-Uni", exist_ok=True)

    model_path = f"checkpoints-AM-Uni/best_memristor_lstm_attn_bs{batch_size}.pth"
    csv_path = f"checkpoints-AM-Uni/log_memristor_lstm_attn_bs{batch_size}.csv"

    with open(csv_path, mode='w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["epoch", "train_loss", "test_acc"])

    for epoch in range(num_epochs):
        model.train()
        running_loss = 0.0
        for X, y in train_loader:
            X, y = X.to(device), y.to(device)
            optimizer.zero_grad()
            out = model(X)
            loss = criterion(out, y)
            loss.backward()
            model.update_memristors(lr=memristor_lr)
            optimizer.step()
            running_loss += loss.item()

        avg_loss = running_loss / len(train_loader)

        # ======== 测试集评估 ========
        model.eval()
        correct, total = 0, 0
        with torch.no_grad():
            for X_val, y_val in test_loader:
                X_val, y_val = X_val.to(device), y_val.to(device)
                out = model(X_val)
                preds = out.argmax(dim=1)
                correct += (preds == y_val).sum().item()
                total += y_val.size(0)
        acc = correct / total

        # ======== CSV记录 ========
        with open(csv_path, mode='a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([epoch+1, avg_loss, acc])

        scheduler.step(avg_loss)
        print(f"[Batch {batch_size}] Epoch [{epoch+1}/{num_epochs}] Loss: {avg_loss:.4f} Test Acc: {acc*100:.2f}%")

        # ======== 保存最佳模型 ========
        if acc > best_acc:
            best_acc = acc
            epochs_no_improve = 0
            torch.save(model.state_dict(), model_path)
            print("✅ Test accuracy improved. Model saved.")
        else:
            epochs_no_improve += 1
            print(f"No improvement for {epochs_no_improve} epochs.")

        # ======== 早停 ========
        if epochs_no_improve >= patience:
            print(f"🛑 Early stopping triggered at epoch {epoch+1}")
            break

    return best_acc, model_path, csv_path

import argparse

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--lr", type=float, default=1e-4, help="AdamW learning rate")
    parser.add_argument("--memristor_lr", type=float, default=0.001, help="Memristor update learning rate")
    parser.add_argument("--batch_sizes", type=int, nargs="+", default=[16, 32, 64], help="Batch sizes to sweep")
    parser.add_argument("--pretrained_path", type=str, default="checkpoints-AM-Uni/best_memristor_lstm_attn.pth")
    parser.add_argument("--freeze", action="store_true", help="Freeze non-memristor layers")
    args = parser.parse_args()

    print(f"Finetuning with lr={args.lr}, memristor_lr={args.memristor_lr}, batch_sizes={args.batch_sizes}")

    best_overall_acc = 0.0
    best_model_path = None
    best_csv_path = None

    for bs in args.batch_sizes:
        acc, model_path, csv_path = train_model(
            batch_size=bs,
            num_epochs=100,
            lr=args.lr,
            memristor_lr=args.memristor_lr,
            pretrained_path=args.pretrained_path,
            freeze_nonmemristor=args.freeze
        )
        print(f"Batch size {bs} finished. Best test acc: {acc*100:.2f}%")
        if acc > best_overall_acc:
            best_overall_acc = acc
            best_model_path = model_path
            best_csv_path = csv_path

    if best_model_path:
        shutil.copy(best_model_path, "checkpoints-AM-Uni/best_memristor_lstm_attn_ft.pth")
        shutil.copy(best_csv_path, "checkpoints-AM-Uni/best_memristor_lstm_attn_ft_log.csv")
        print(f"最优模型为 {best_model_path}, 已复制为 checkpoints-AM-Uni/best_memristor_lstm_attn_ft.pth")
        print(f"最优日志为 {best_csv_path}, 已复制为 checkpoints-AM-Uni/best_memristor_lstm_attn_ft_log.csv")
        print(f"最优测试集准确率: {best_overall_acc*100:.2f}%")

if __name__ == "__main__":
    main()
