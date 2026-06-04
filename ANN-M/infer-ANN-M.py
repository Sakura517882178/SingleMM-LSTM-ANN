import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader
import numpy as np
import csv
import openpyxl

# ===========================
# 从 Y6 实验数据提取忆阻器参数 (log 域)
# ===========================
def load_y6_params(xlsx_path="Y6-data.xlsx", stable_after_ms=240.0):
    """从 Y6 单分子结 I-V 数据提取 log 域电导参数。
    Light: 50Hz 光脉冲调制，取稳定区 P5/P99.9 作为动态范围。
    Dark:  无光调制，电导应为定值; 滤除仪器掉坑(G > median-1.5)后取均值。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    time, voltage, cond = [], [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            time.append(float(row[0]))
        if row[1] is not None:
            voltage.append(float(row[1]))
        if row[3] is not None:
            cond.append(float(row[3]))
    t_arr = np.array(time)
    v_arr = np.array(voltage)
    c_arr = np.array(cond)
    stable_mask = t_arr > stable_after_ms
    c_stable = c_arr[stable_mask]
    V_exp = float(np.max(v_arr))
    median = float(np.median(c_stable))
    tail_frac = np.mean(c_stable < median - 1.5)
    if tail_frac > 0.05:
        c_clean = c_stable[c_stable > -5]
        G_val = float(np.mean(c_clean))
        return G_val, G_val, V_exp
    else:
        return float(np.percentile(c_stable, 5)), float(np.percentile(c_stable, 99.9)), V_exp

# ===========================
# Dataset
# ===========================
class SpeechFeatureDataset(torch.utils.data.Dataset):
    def __init__(self, npz_file):
        data = np.load(npz_file, allow_pickle=True)
        self.X = data["X"]
        self.y = data["y"]
        if self.y.min() < 0 or self.y.max() >= len(np.unique(self.y)):
            raise ValueError(f"标签范围不正确: min={self.y.min()}, max={self.y.max()}")

    def __len__(self):
        return len(self.X)

    def __getitem__(self, idx):
        return torch.tensor(self.X[idx], dtype=torch.float32), torch.tensor(self.y[idx], dtype=torch.long)

def collate_fn(batch):
    feats, labels = zip(*batch)
    return torch.stack(feats, dim=0), torch.stack(labels, dim=0)

# ===========================
# 忆阻器线性层（log 域: I = 10^G * V）
# ===========================
class MemristorLinearI(nn.Module):
    def __init__(self, in_features, out_features, Gmin_log, Gmax_log, V, step=0.01, noise=0.0):
        super().__init__()
        self.Gmin = Gmin_log
        self.Gmax = Gmax_log
        self.step = step
        self.noise = noise
        self.V = V
        self.G = nn.Parameter(torch.empty(out_features, in_features).uniform_(Gmin_log, Gmax_log))

    def forward(self, x):
        I = 10 ** self.G * self.V
        return F.linear(x, I)

# ===========================
# 忆阻器非线性层（log 域: I = x * 10^G * V）
# ===========================
class MemristorNonlinear(nn.Module):
    def __init__(self, Gmin_log, Gmax_log, V, step=0.01, noise=0.0):
        super().__init__()
        self.V = V
        self.Gmin = Gmin_log
        self.Gmax = Gmax_log
        self.step = step
        self.noise = noise
        self.G = nn.Parameter(torch.empty(1).uniform_(Gmin_log, Gmax_log))

    def forward(self, x):
        x = torch.clamp(x, min=0.0)
        I = x * (10 ** self.G * self.V)
        return I

# ===========================
# 忆阻器 ANN + LSTM + 注意力
# ===========================
class MemristorANN_LSTM_Attn(nn.Module):
    def __init__(self, input_dim=64, hidden_dim=128, lstm_layers=1,
                 fc_hidden=128, num_classes=36, dropout=0.3,
                 Gmin_log=-7.92, Gmax_log=-2.81, V=0.35):
        super().__init__()
        self.hidden_dim = hidden_dim

        self.lstm = nn.LSTM(input_dim, hidden_dim, lstm_layers,
                            batch_first=True, bidirectional=False)

        self.attn_weight = nn.Parameter(torch.Tensor(hidden_dim, 1))
        nn.init.xavier_uniform_(self.attn_weight)

        self.fc1 = MemristorLinearI(hidden_dim, fc_hidden, Gmin_log, Gmax_log, V)
        self.layernorm = nn.LayerNorm(fc_hidden)
        self.dropout = nn.Dropout(dropout)
        self.nonlinear = MemristorNonlinear(Gmin_log, Gmax_log, V)
        self.fc2 = MemristorLinearI(fc_hidden, num_classes, Gmin_log, Gmax_log, V)

    def forward(self, x):
        lstm_out, _ = self.lstm(x)
        attn_scores = torch.tanh(lstm_out @ self.attn_weight)
        attn_weights = torch.softmax(attn_scores, dim=1)
        pooled = torch.sum(lstm_out * attn_weights, dim=1)

        x = self.nonlinear(self.fc1(pooled))
        x = self.layernorm(x)
        x = self.dropout(x)
        logits = self.fc2(x)
        return logits

    def forward(self, x):
        lstm_out, _ = self.lstm(x)
        attn_scores = torch.tanh(lstm_out @ self.attn_weight)
        attn_weights = torch.softmax(attn_scores, dim=1)
        pooled = torch.sum(lstm_out * attn_weights, dim=1)

        x = self.nonlinear(self.fc1(pooled))
        x = self.layernorm(x)
        x = self.dropout(x)
        logits = self.fc2(x)
        return logits

# ===========================
# 推理函数
# ===========================
def inference(npz_file, model_path, device=None, batch_size=16):
    device = device or ("cuda" if torch.cuda.is_available() else "cpu")

    dataset = SpeechFeatureDataset(npz_file)
    loader = DataLoader(dataset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)

    # 加载 Y6 参数
    Gmin_log, Gmax_log, V_exp = load_y6_params("data.xlsx")

    num_classes = len(np.unique(dataset.y))
    model = MemristorANN_LSTM_Attn(input_dim=64, hidden_dim=128, fc_hidden=128, num_classes=num_classes,
                                   Gmin_log=Gmin_log, Gmax_log=Gmax_log, V=V_exp)
    model.load_state_dict(torch.load(model_path, map_location=device))
    model.to(device)
    model.eval()
    
    all_preds = []
    all_labels = []
    
    with torch.no_grad():
        for X, y in loader:
            X = X.to(device)
            logits = model(X)
            preds = torch.argmax(logits, dim=1).cpu().numpy()
            all_preds.append(preds)
            all_labels.append(y.numpy())
    
    all_preds = np.concatenate(all_preds)
    all_labels = np.concatenate(all_labels)
    accuracy = (all_preds == all_labels).mean()
    print(f"✅ 推理完成。准确率: {accuracy*100:.2f}%")
    
    return all_preds, all_labels, accuracy

# ===========================
# 主函数示例
# ===========================

import argparse

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--test_npz", type=str, default="test_features.npz")
    parser.add_argument("--model", type=str, default="checkpoints-AM-Uni/best_memristor_lstm_attn_ft.pth")
    parser.add_argument("--output", type=str, default="predictions-ANN-M.csv")
    args = parser.parse_args()

    preds, labels, acc = inference(args.test_npz, args.model)

    with open(args.output, mode="w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["true", "pred"])
        for y, p in zip(labels, preds):
            writer.writerow([y, p])
    print(f"Inference accuracy: {acc*100:.2f}%")
    print(f"Predictions saved to {args.output}")

if __name__ == "__main__":
    main()
