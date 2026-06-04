import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader
import numpy as np
import os
import shutil
import csv
import openpyxl

# ===========================
# 从 Y6 实验数据提取忆阻器参数 (log 域)
# ===========================
def load_y6_params(xlsx_path="Y6-data.xlsx", stable_after_ms=240.0):
    """从 Y6 单分子结 I-V 数据提取 log 域电导参数。
    Light: 50Hz 光脉冲调制，取稳定区 P5/P99.9 作为动态范围。
    Dark:  无光调制，电导应为定值; 滤除仪器掉坑(G > median-1.5)后取均值。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    time, voltage, cond = [], [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            time.append(float(row[0]))
        if row[1] is not None:
            voltage.append(float(row[1]))
        if row[3] is not None:
            cond.append(float(row[3]))
    t_arr = np.array(time)
    v_arr = np.array(voltage)
    c_arr = np.array(cond)
    stable_mask = t_arr > stable_after_ms
    c_stable = c_arr[stable_mask]
    V_exp = float(np.max(v_arr))

    # 判断暗态/光态: 暗态有大量仪器掉坑点远离主峰
    median = float(np.median(c_stable))
    tail_frac = np.mean(c_stable < median - 1.5)
    is_dark = tail_frac > 0.05

    if is_dark:
        # 暗态: 滤除仪器掉坑 (G > -5), 取稳定平台均值作为固定电导
        c_clean = c_stable[c_stable > -5]
        G_val = float(np.mean(c_clean))
        Gmin_log = G_val
        Gmax_log = G_val
        print(f"[Y6-Dark] stable t>{stable_after_ms}ms, filtered {100*tail_frac:.1f}% artifacts, "
              f"G={G_val:.4f} (log G/G0, fixed), V={V_exp:.2f} V")
    else:
        # 光态: 取稳定区 P5/P99.9 作为忆阻器工作范围
        Gmin_log = float(np.percentile(c_stable, 5))
        Gmax_log = float(np.percentile(c_stable, 99.9))
        print(f"[Y6-Light] stable t>{stable_after_ms}ms, P5/P99.9, "
              f"Gmin={Gmin_log:.4f}, Gmax={Gmax_log:.4f} (log G/G0), V={V_exp:.2f} V")

    return Gmin_log, Gmax_log, V_exp

# ===========================
# Dataset
# ===========================
class SpeechFeatureDataset(torch.utils.data.Dataset):
    def __init__(self, npz_file):
        data = np.load(npz_file, allow_pickle=True)
        self.X = data["X"]
        self.y = data["y"]
        if self.y.min() < 0 or self.y.max() >= len(np.unique(self.y)):
            raise ValueError(f"标签范围不正确: min={self.y.min()}, max={self.y.max()}")

    def __len__(self):
        return len(self.X)

    def __getitem__(self, idx):
        return torch.tensor(self.X[idx], dtype=torch.float32), torch.tensor(self.y[idx], dtype=torch.long)

def collate_fn(batch):
    feats, labels = zip(*batch)
    return torch.stack(feats, dim=0), torch.stack(labels, dim=0)

# ===========================
# 忆阻器线性层（log 域电导: I = 10^G * V）
# G 存储在 log(G/G0) 域，范围来自 Y6 实验数据
# ===========================
class MemristorLinearI(nn.Module):
    def __init__(self, in_features, out_features,
                 Gmin_log, Gmax_log, V, step=0.01, noise=0.0):
        super().__init__()
        self.Gmin = Gmin_log
        self.Gmax = Gmax_log
        self.step = step
        self.noise = noise
        self.V = V
        self.G = nn.Parameter(torch.empty(out_features, in_features).uniform_(Gmin_log, Gmax_log))

    def forward(self, x):
        I = 10 ** self.G * self.V
        return F.linear(x, I)

    def update_memristor(self, lr=0.01):
        if self.G.grad is not None:
            delta = -lr * self.G.grad.data
            delta = torch.clamp(delta, -self.step, self.step)
            self.G.data += delta
            if self.noise > 0:
                self.G.data += self.noise * torch.randn_like(self.G)
            self.G.data = torch.clamp(self.G.data, self.Gmin, self.Gmax)
            self.G.grad.zero_()

# ===========================
# 忆阻器非线性层（log 域电导，替代 ReLU）
# G 存储在 log(G/G0) 域，范围来自 Y6 实验数据
# ===========================
class MemristorNonlinear(nn.Module):
    def __init__(self, Gmin_log, Gmax_log, V, step=0.01, noise=0.0):
        super().__init__()
        self.V = V
        self.Gmin = Gmin_log
        self.Gmax = Gmax_log
        self.step = step
        self.noise = noise
        self.G = nn.Parameter(torch.empty(1).uniform_(Gmin_log, Gmax_log))

    def forward(self, x):
        x = torch.clamp(x, min=0.0)
        I = x * (10 ** self.G * self.V)
        return I

    def update_memristor(self, lr=0.001):
        if self.G.grad is not None:
            delta = -lr * self.G.grad.data
            delta = torch.clamp(delta, -self.step, self.step)
            self.G.data += delta
            if self.noise > 0:
                self.G.data += self.noise * torch.randn_like(self.G)
            self.G.data = torch.clamp(self.G.data, self.Gmin, self.Gmax)
            self.G.grad.zero_()

# ===========================
# 忆阻器版 ANN + 单向 LSTM + 可训练注意力
# ===========================
class MemristorANN_LSTM_Attn(nn.Module):
    def __init__(self, input_dim=64, hidden_dim=128, lstm_layers=1,
                 fc_hidden=128, num_classes=36, dropout=0.3,
                 Gmin_log=-7.92, Gmax_log=-2.81, V=0.35):
        super().__init__()
        self.hidden_dim = hidden_dim
        self.num_directions = 1  # 单向 LSTM

        self.lstm = nn.LSTM(input_dim, hidden_dim, lstm_layers,
                            batch_first=True, bidirectional=False)

        # 可训练注意力
        self.attn_weight = nn.Parameter(torch.Tensor(hidden_dim, 1))
        nn.init.xavier_uniform_(self.attn_weight)

        # 忆阻器 FC + Dropout + LayerNorm (log 域 G, 参数来自 Y6 实验)
        self.fc1 = MemristorLinearI(hidden_dim, fc_hidden, Gmin_log, Gmax_log, V)
        self.layernorm = nn.LayerNorm(fc_hidden)
        self.dropout = nn.Dropout(dropout)
        self.nonlinear = MemristorNonlinear(Gmin_log, Gmax_log, V)
        self.fc2 = MemristorLinearI(fc_hidden, num_classes, Gmin_log, Gmax_log, V)

    def forward(self, x):
        lstm_out, _ = self.lstm(x)                      # [B, T, hidden_dim]
        attn_scores = torch.tanh(lstm_out @ self.attn_weight)  # [B, T, 1]
        attn_weights = torch.softmax(attn_scores, dim=1)       # [B, T, 1]
        pooled = torch.sum(lstm_out * attn_weights, dim=1)     # [B, hidden_dim]

        x = self.nonlinear(self.fc1(pooled))
        x = self.layernorm(x)
        x = self.dropout(x)
        logits = self.fc2(x)
        return logits

    def update_memristors(self, lr=0.01):
        for m in self.modules():
            if isinstance(m, MemristorLinearI) or isinstance(m, MemristorNonlinear):
                m.update_memristor(lr)

# ===========================
# 训练函数
# ===========================
def train_model(batch_size, num_epochs=200, patience=30, lr=0.003, memristor_lr=0.001,
                y6_data="Y6-data.xlsx", tag=""):
    device = "cuda" if torch.cuda.is_available() else "cpu"

    # 从 Y6 实验数据读取忆阻器参数
    Gmin_log, Gmax_log, V_exp = load_y6_params(y6_data)
    tag_str = f"_{tag}" if tag else ""

    train_dataset = SpeechFeatureDataset("train_features.npz")
    test_dataset = SpeechFeatureDataset("test_features.npz")
    train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True, collate_fn=collate_fn)
    test_loader = DataLoader(test_dataset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)

    num_classes = len(np.unique(train_dataset.y))
    model = MemristorANN_LSTM_Attn(input_dim=64, hidden_dim=128, lstm_layers=1,
                                   fc_hidden=128, num_classes=num_classes,
                                   Gmin_log=Gmin_log, Gmax_log=Gmax_log, V=V_exp).to(device)

    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.AdamW(model.parameters(), lr=lr, weight_decay=1e-5)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.5, patience=5)

    best_acc = 0.0
    epochs_no_improve = 0
    os.makedirs("checkpoints-AM-Uni", exist_ok=True)

    model_path = f"checkpoints-AM-Uni/best_memristor_lstm_attn_bs{batch_size}{tag_str}.pth"
    csv_path = f"checkpoints-AM-Uni/log_memristor_lstm_attn_bs{batch_size}{tag_str}.csv"

    with open(csv_path, mode='w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["epoch", "train_loss", "test_acc"])

    for epoch in range(num_epochs):
        model.train()
        running_loss = 0.0
        for X, y in train_loader:
            X, y = X.to(device), y.to(device)
            optimizer.zero_grad()
            out = model(X)
            loss = criterion(out, y)
            loss.backward()
            model.update_memristors(lr=memristor_lr)
            optimizer.step()
            running_loss += loss.item()

        avg_loss = running_loss / len(train_loader)

        model.eval()
        correct, total = 0, 0
        with torch.no_grad():
            for X_val, y_val in test_loader:
                X_val, y_val = X_val.to(device), y_val.to(device)
                out = model(X_val)
                preds = out.argmax(dim=1)
                correct += (preds == y_val).sum().item()
                total += y_val.size(0)
        acc = correct / total

        with open(csv_path, mode='a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([epoch+1, avg_loss, acc])

        scheduler.step(avg_loss)
        print(f"[Batch {batch_size}] Epoch [{epoch+1}/{num_epochs}] Loss: {avg_loss:.4f} Test Acc: {acc*100:.2f}%")

        if acc > best_acc:
            best_acc = acc
            epochs_no_improve = 0
            torch.save(model.state_dict(), model_path)
            print("Test accuracy improved. Model saved.")
        else:
            epochs_no_improve += 1
            print(f"No improvement for {epochs_no_improve} epochs.")

        if epochs_no_improve >= patience:
            print(f"Early stopping triggered at epoch {epoch+1}")
            break

    return best_acc, model_path, csv_path

import argparse

# ===========================
# 主函数
# ===========================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--lr", type=float, default=0.003, help="AdamW learning rate")
    parser.add_argument("--memristor_lr", type=float, default=0.001, help="Memristor update learning rate")
    parser.add_argument("--batch_sizes", type=int, nargs="+", default=[16, 32, 64], help="Batch sizes to sweep")
    parser.add_argument("--y6_data", type=str, default="Y6-data.xlsx", help="Y6 I-V data file (Light/Dark control)")
    parser.add_argument("--tag", type=str, default="", help="Tag for checkpoint naming")
    args = parser.parse_args()

    print(f"Training with y6_data={args.y6_data}, lr={args.lr}, memristor_lr={args.memristor_lr}, batch_sizes={args.batch_sizes}")

    best_overall_acc = 0.0
    best_model_path = None
    best_csv_path = None

    for bs in args.batch_sizes:
        acc, model_path, csv_path = train_model(
            batch_size=bs,
            lr=args.lr,
            memristor_lr=args.memristor_lr,
            y6_data=args.y6_data,
            tag=args.tag
        )
        print(f"Batch size {bs} finished. Best test acc: {acc*100:.2f}%")
        if acc > best_overall_acc:
            best_overall_acc = acc
            best_model_path = model_path
            best_csv_path = csv_path

    if best_model_path:
        tag_str = f"_{args.tag}" if args.tag else ""
        shutil.copy(best_model_path, f"checkpoints-AM-Uni/best_memristor_lstm_attn{tag_str}.pth")
        shutil.copy(best_csv_path, f"checkpoints-AM-Uni/best_memristor_lstm_attn{tag_str}_log.csv")
        print(f"最优模型为 {best_model_path}, 已复制为 checkpoints-AM-Uni/best_memristor_lstm_attn{tag_str}.pth")
        print(f"最优训练日志为 {best_csv_path}, 已复制为 checkpoints-AM-Uni/best_memristor_lstm_attn{tag_str}_log.csv")
        print(f"最优测试集准确率: {best_overall_acc*100:.2f}%")

if __name__ == "__main__":
    main()
